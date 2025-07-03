# your_app/management/commands/limpiar_archivos.py

import os
import datetime
import shutil
from django.core.management.base import BaseCommand
from django.conf import settings
from django.apps import apps
from openpyxl import Workbook

from myapp.models import (
    Soat,
    RevisionTecnomecanica,
    TarjetaOperacion,
    PolizaContractual,
    PolizaExtracontractual,
    Vehiculos,
    EventoDocumento,
    EventoLegalFile,
    EventEvidence,
    SiniestroHistorialActuacion,
    DocumentoSiniestroHistorialActuacion,
    Tercero,
    ActaConciliacion,
    HistorialActuacion,
)

EXCLUDED_DIRS = {'assets', 'contratos', 'pdf_eventos'}
SKIP_PREFIXES = ('~$', 'reporte_archivos_')

class Command(BaseCommand):
    help = 'Reporte, backup y limpieza de archivos en MEDIA_ROOT no referenciados en la BD.'

    def handle(self, *args, **options):
        media_root = settings.MEDIA_ROOT
        referenced = set()

        # 1. Referencias en FileField / ImageField
        for model in apps.get_models():
            for field in model._meta.get_fields():
                if hasattr(field, 'upload_to') and getattr(field, 'upload_to'):
                    for obj in model.objects.all():
                        f = getattr(obj, field.name)
                        if f:
                            name = f.name if hasattr(f, 'name') else f
                            if name:
                                referenced.add(name)

        # 2. Referencias en CharField que guardan rutas
        char_models = [
            (Soat, 'soporte'),
            (RevisionTecnomecanica, 'soporte'),
            (TarjetaOperacion, 'soporte'),
            (PolizaContractual, 'soporte'),
            (PolizaExtracontractual, 'soporte'),
        ]
        for model, field_name in char_models:
            for obj in model.objects.all():
                val = getattr(obj, field_name, '') or ''
                if val and not val.startswith('http'):
                    referenced.add(val)

        # 3. Archivos reales en MEDIA_ROOT (sin backups previos)
        actual = set()
        for root, _, files in os.walk(media_root):
            for filename in files:
                rel = os.path.relpath(os.path.join(root, filename), media_root).replace('\\', '/')
                if rel.startswith('backup_eliminados/'):
                    continue
                actual.add(rel)

        mantenidos = sorted(actual & referenced)

        unreferenced = [
            f for f in actual - referenced
            if not any(os.path.basename(f).startswith(pref) for pref in SKIP_PREFIXES)
        ]

        saltados = [f for f in unreferenced if f.split('/', 1)[0] in EXCLUDED_DIRS]
        eliminar = [f for f in unreferenced if f.split('/', 1)[0] not in EXCLUDED_DIRS]

        # 4. Reporte en consola
        self.stdout.write('\nArchivos que se MANTIENEN:')
        for f in mantenidos:
            self.stdout.write(f'  {f}')

        self.stdout.write('\nArchivos SIN REFERENCIA (excluidos de eliminación):')
        for f in saltados:
            self.stdout.write(f'  {f}')

        self.stdout.write('\nArchivos que se ELIMINARÁN (backup):')
        for f in eliminar:
            self.stdout.write(f'  {f}')

        # 5. Generar reporte Excel
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = f'reporte_archivos_{timestamp}.xlsx'
        report_path = os.path.join(media_root, report_name)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Mantenidos"
        ws1.append(["Archivo"])
        for f in mantenidos:
            ws1.append([f])

        ws2 = wb.create_sheet(title="Saltados")
        ws2.append(["Archivo"])
        for f in saltados:
            ws2.append([f])

        ws3 = wb.create_sheet(title="A_Eliminar")
        ws3.append(["Archivo"])
        for f in eliminar:
            ws3.append([f])

        wb.save(report_path)
        self.stdout.write(f'\nReporte generado: {report_path}')

        # 6. Confirmación de movimiento a backup
        if not eliminar:
            self.stdout.write('\nNo hay archivos a eliminar.')
            return

        confirm = input('\n¿Mover a backup estos archivos? (s/n): ').strip().lower()
        if confirm != 's':
            self.stdout.write('Operación cancelada.')
            return

        backup_root = os.path.join(media_root, f'backup_eliminados/{timestamp}')
        moved = []
        for f in eliminar:
            src = os.path.join(media_root, f)
            if not os.path.exists(src):
                self.stderr.write(f'Skip (no existe): {f}')
                continue
            dst = os.path.join(backup_root, f)
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            try:
                shutil.move(src, dst)
                self.stdout.write(f'Movido a backup: {f}')
                moved.append(f)
            except Exception as e:
                self.stderr.write(f'Error moviendo {f}: {e}')

        self.stdout.write(f'\nBackup en: {backup_root}')

        # 7. Validación: revertir o confirmar
        action = input('\nRevise el backup. Elija: revertir / confirmar / cancelar: ').strip().lower()
        if action == 'revertir':
            for f in moved:
                bsrc = os.path.join(backup_root, f)
                bdst = os.path.join(media_root, f)
                if not os.path.exists(bsrc):
                    self.stderr.write(f'Falta en backup (no revertido): {f}')
                    continue
                os.makedirs(os.path.dirname(bdst), exist_ok=True)
                try:
                    shutil.move(bsrc, bdst)
                    self.stdout.write(f'Revertido: {f}')
                except Exception as e:
                    self.stderr.write(f'Error revirtiendo {f}: {e}')
            self.stdout.write('\nTodos los archivos revertidos.')
        elif action == 'confirmar':
            try:
                shutil.rmtree(backup_root)
                self.stdout.write('\nBackup eliminado. Operación finalizada.')
            except Exception as e:
                self.stderr.write(f'Error eliminando backup: {e}')
        else:
            self.stdout.write('\nSin acción adicional. Archivos permanecen en backup.')
