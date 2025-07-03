from .AdminTriggerXlsx import write_name_company, write_month_year, process_data
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook
from datetime import datetime
import json
import os
import io


def generate_excel_response(workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Excel.xlsx'
    return response


@csrf_exempt
@require_http_methods(["POST"])
def XlsxRP_CuotaAdmon(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            Opcion = int(results['tipoInforme'])
            empresa = int(results['empresa'])
            try:
                startDate = results['startDate']
                startDate = datetime.strptime(
                    startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                year = startDate.year
                month = startDate.month
            except:
                pass

            script_dir = os.path.dirname(__file__)

            if Opcion == 2:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Contabilidad_RptoCuotaAdmon_Ciudades.xlsx')
            elif Opcion == 3:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Contabilidad_RptoCuotaAdmon_Propietarios.xlsx')
            elif Opcion == 4:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Contabilidad_RptoCuotaAdmon_Ciudades-Colibertador.xlsx')
            elif Opcion == 5:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Contabilidad_RptoCuotaAdmon_AgenciasBerlitur.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                # Cargar la plantilla Excel
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                try:
                    write_month_year(sheet, month, year, 'I', 6, 'G', 6)
                except:
                    pass

                # Ubicacion para nombre de empresa
                if Opcion != 12 and Opcion != 14 and Opcion != 28 and Opcion != 29 and Opcion != 31 and Opcion != 9 and Opcion != 34:
                    start_column_colEmp = 'D'
                    start_row_colEmp = 2
                else:
                    start_column_colEmp = 'C'
                    start_row_colEmp = 2

                write_name_company(
                    sheet, empresa, column_colEmp=start_column_colEmp, row_colEmp=start_row_colEmp)

                if Opcion == 2:
                    process_data(sheet, datos, 8, columns_to_sum=[
                                 4, 5], columns_to_average=[0])
                elif Opcion == 3:
                    process_data(sheet, datos, 8, columns_to_sum=[
                                 6, 7], columns_to_average=[0])
                elif Opcion == 4:
                    process_data(sheet, datos, 8, columns_to_sum=[
                                 5, 6], columns_to_average=[0])
                elif Opcion == 5:
                    process_data(sheet, datos, 8, columns_to_sum=[
                                 3, 4], columns_to_average=[0])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxSp_RptHistoFuec(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            empresa = results['empresa']
            startDate = results['startDate']
            startDate = datetime.strptime(
                startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
            year = startDate.year
            month = startDate.month

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            plantilla_path = os.path.join(
                script_dir, '../docs/Plantillas/Plantilla_Rpto_Otros_InformeFuec.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                process_data(sheet, datos, 8, columns_to_sum=[],
                             columns_to_average=[])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxSp_RP_Prueba_Alcoholimetria(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            empresa = results['empresa']
            startDate = results['startDate']
            startDate = datetime.strptime(
                startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
            year = startDate.year
            month = startDate.month

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            plantilla_path = os.path.join(
                script_dir, '../docs/Plantillas/Plantilla_Rpto_Alcoholimetria_Detallado.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                process_data(sheet, datos, 8, columns_to_sum=[],
                             columns_to_average=[])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxRP_consultas01(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results["results"])
            Opcion = int(results['Opcion'])
            empresa = int(results['empresa'])
            if Opcion != 12 and Opcion != 99:
                startDate = results['startDate']
                startDate = datetime.strptime(
                    startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                year = startDate.year
                month = startDate.month
            try:
                SubOpcion = int(results['SubOpcion'])
            except:
                SubOpcion = None
            try:
                datos = json.loads(results.get("results", []))
                datos2 = json.loads(results.get("results2", []))
            except:
                pass

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            if Opcion == 6:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Boleterias_Estadisticas_PlanilladosPasajerosXRutas.xlsx')
            elif Opcion == 7:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Boleterias_Estadisticas_PlanilladosPasajerosXTrayectos.xlsx')
            elif Opcion == 8:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Boleterias_Estadisticas_PlanilladosPasajerosOrigen-Destino.xlsx')
            elif Opcion == 9:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Boleterias_Estadisticas_TiquetesPromedioXBoleteria.xlsx')
            elif Opcion == 16:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_TesoreriaComisionGuillermoCifuentes.xlsx')
            elif Opcion == 22:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Boleterias_Estadisticas_PlanilladosPasajerosXTrayectos.xlsx')
            elif Opcion == 33:
                if SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_TesoreriaTiquetesMacarena.xlsx')
            elif Opcion == 99:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_Certificaciones.xlsx')

            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook['Sheet1']

                try:
                    write_month_year(sheet, month, year, 'I', 6, 'G', 6)
                except:
                    pass

                # Ubicacion para nombre de empresa
                if Opcion != 12 and Opcion != 14 and Opcion != 28 and Opcion != 29 and Opcion != 31 and Opcion != 9 and Opcion != 34:
                    start_column_colEmp = 'D'
                    start_row_colEmp = 2
                else:
                    start_column_colEmp = 'C'
                    start_row_colEmp = 2

                try:
                    write_name_company(
                        sheet, empresa, column_colEmp=start_column_colEmp, row_colEmp=start_row_colEmp)
                except:
                    pass

                if Opcion == 0:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        4, 6, 7, 8, 9], columns_to_average=[0])
                elif Opcion == 6:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        3, 4, 5, 6], columns_to_average=[0])
                elif Opcion == 7:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        4, 5], columns_to_average=[0])
                elif Opcion == 8:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        3, 4, 7, 8, 9, 10], columns_to_average=[9, 10])
                elif Opcion == 9:
                    try:
                        process_data(sheet, datos[0], 8, columns_to_sum=[
                            6, 7, 8], columns_to_average=[9, 10])
                        process_data(workbook["Consolidado"], datos[1], 8, columns_to_sum=[
                            4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15], columns_to_average=[9, 10])
                    except:
                        pass
                elif Opcion == 22:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        0], columns_to_average=[0])
                elif Opcion == 99:
                    process_data(sheet, datos, 8, columns_to_sum=[
                        0], columns_to_average=[0])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlxsRP_Macarena(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            empresa = results['empresa']
            Opcion = int(results['Opcion'])
            startDate = results['startDate']
            startDate = datetime.strptime(
                startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
            year = startDate.year
            month = startDate.month

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            if Opcion == 1:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Otros_ConvenioMacarena_RelacionDeTiquetes.xlsx')
            elif Opcion == 2:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Otros_ConvenioMacarena_RelacionDeTiquetesVencidos.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                process_data(sheet, datos, start_row_col=8,
                             columns_to_sum=[0], columns_to_average=[])
                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxFics_MicroSegurosGET(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = results['results']
            empresa = int(results['empresa'])
            startDate = results['startDate']
            startDate = datetime.strptime(startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
            year = startDate.year
            month = startDate.month

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            plantilla_path = os.path.join(
                script_dir, '../docs/Plantillas/Plantilla_Rpto_MicroSeguros_TiquetesVendidosConMicroSeguros.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                process_data(sheet, datos, 8, columns_to_sum=[2, 3, 4, 5, 6],
                             columns_to_average=[])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxRP_MIGRACION(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            empresa = int(results['empresa'])
            Opcion = int(results['Opcion'])
            startDate = results['startDate']
            try:
                year = startDate.year
                month = startDate.month
            except:
                pass

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            if Opcion == 1:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_MacarenaTiquetes.xlsx')
            elif Opcion == 2:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_Migracion_RptoMigracionCXFecha.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                try:
                    write_month_year(sheet, month, year, column_colMes,
                                     row_colMes, column_colYear, row_colYear)
                except:
                    pass

                process_data(sheet, datos, 8, columns_to_sum=[],
                             columns_to_average=[])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxRPT_EstadisticaXTaquilla(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            month = results['Month']
            year = results['Year']
            Opcion = int(results['Opcion'])

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)
            if Opcion == 0:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosDeVentaVentasTaquillerosBogota.xlsx')
            elif Opcion == 1:
                plantilla_path = os.path.join(
                    script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosDeVentaVentasTaquillerosBogotaFecha.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                process_data(sheet, datos, 8, columns_to_sum=[
                             0], columns_to_average=[0])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@require_http_methods(["POST"])
def XlsxRP_Consultas05(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            Opcion = int(results['Opcion'])
            SubOpcion = int(results['SubOpcion'])
            empresa = int(results['empresa'])
            print(Opcion, SubOpcion)
            try:
                startDate = results['startDate']
                startDate = datetime.strptime(
                    startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
                year = startDate.year
                month = startDate.month
            except:
                pass

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            if Opcion == 0:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_Viajes.xlsx')
            elif Opcion == 1:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_InformePaxMovilizadosConsolidadoTotal.xlsx')
                elif SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_InformePaxMovilizadosConsolidadoXLinea.xlsx')
                elif SubOpcion == 2:
                    if empresa == 277 or empresa == 278:
                        print("Paso1")
                        plantilla_path = os.path.join(
                            script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_BusquedaPorFechas_277-278.xlsx')
                    elif empresa == 9001 or empresa == 310 or empresa == 320:
                        print("Paso2")
                        plantilla_path = os.path.join(
                            script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_BusquedaPorFechasElse.xlsx')
            if Opcion == 9:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Planeacion_OcupacionXLineas_Consolidado.xlsx')
                if SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Planeacion_OcupacionXLineas_Detallado.xlsx')
                if SubOpcion == 2:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Planeacion_OcupacionXLineas_ConsolidadoPorRuta.xlsx')
                if SubOpcion == 3:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Planeacion_OcupacionXLineas_ConsolidadoMensual.xlsx')
                if SubOpcion == 4:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Planeacion_OcupacionXLineas_ConsolidadoPorHorario.xlsx')
            elif Opcion == 11:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_DomiciliosPorLinea.xlsx')
            elif Opcion == 12:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_ViajesXPersona.xlsx')
            elif Opcion == 13:
                if SubOpcion == 0:
                    try:
                        SubOpcion2 = int(results['SubOpcion2'])
                        if SubOpcion2 == 0:
                            plantilla_path = os.path.join(
                                script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosVenta_PaxVendidosCucuta.xlsx')
                        elif SubOpcion2 == 1:
                            plantilla_path = os.path.join(
                                script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosDeVentaPaxVendidosBogota.xlsx')
                    except:
                        pass
            elif Opcion == 14:
                if SubOpcion == 0:
                    try:
                        SubOpcion2 = int(results['SubOpcion2'])
                        if SubOpcion2 == 0:
                            plantilla_path = os.path.join(
                                script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosVenta_PaxDespachadosEnBogota.xlsx')
                        if SubOpcion2 == 1:
                            plantilla_path = os.path.join(
                                script_dir, '../docs/Plantillas/Plantilla_Rpto_PuntosVenta_PaxMovilizadosXCiudad.xlsx')
                    except:
                        pass
            elif Opcion == 16:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_TesoreriaComisionGuillermoCifuentes.xlsx')
            elif Opcion == 20:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_PLC_InformeConsolidadoPLC.xlsx')
                elif SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_PLC_InformeDetalladoPLC.xlsx')
            elif Opcion == 26:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_PLEA_InformeConsolidadoPLEA.xlsx')
                elif SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_PLEA_InformeDetalladoPLEA.xlsx')
            elif Opcion == 28:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Comercial_TrazabilidadVentas.xlsx')
            elif Opcion == 29:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_InformeConsolidadoCombustible.xlsx')
                elif SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Operaciones_InformeDetalladoCombustible.xlsx')
            elif Opcion == 30:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_TesoreriaTiquetesMacarena.xlsx')
            elif Opcion == 31:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_MicroSegurosRptoMicroSeguros.xlsx')
            elif Opcion == 33:
                if SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_Biometricos_RptoXFecha.xlsx')

            elif Opcion == 34:
                if SubOpcion == 0:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_NumeroDespachosXDia_Consolidado.xlsx')
                elif SubOpcion == 1:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_NumeroDespachosXDia_Detallado.xlsx')
                elif SubOpcion == 2:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_CantidadVehiculosTrabajaronXDia_Consolidado.xlsx')
                elif SubOpcion == 3:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_CantidadVehiculosTrabajaronXDia_Detallado.xlsx')
                elif SubOpcion == 4:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_HorasEfectivasConductores_Consolidado.xlsx')
                elif SubOpcion == 5:
                    plantilla_path = os.path.join(
                        script_dir, '../docs/Plantillas/Plantilla_Rpto_SeguridadSalud_HorasEfectivasConductores_Detallado.xlsx')

            # Verificar si existe el archivo y intentar cargarlo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                try:
                    write_month_year(sheet, month, year, 'I', 6, 'G', 6)
                except:
                    pass

                # Ubicacion para nombre de empresa
                if Opcion != 12 and Opcion != 14 and Opcion != 28 and Opcion != 29 and Opcion != 31 and Opcion != 9 and Opcion != 34:
                    start_column_colEmp = 'D'
                    start_row_colEmp = 2
                else:
                    start_column_colEmp = 'C'
                    start_row_colEmp = 2

                write_name_company(
                    sheet, empresa, column_colEmp=start_column_colEmp, row_colEmp=start_row_colEmp)

                # Llamar a la función process_data con los parámetros adecuados, proporcionando la lista de columnas a sumar opcionalmente y la lista de columnas a promediar
                if Opcion == 0:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                            4, 6, 7, 8, 9], columns_to_average=[0])
                elif Opcion == 1:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     3, 4, 5, 6, 7], columns_to_average=[7])
                    elif SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4, 5, 6, 7, 8], columns_to_average=[8])
                    elif SubOpcion == 2:
                        if empresa == 277 or empresa == 278:
                            process_data(sheet, datos, 8, columns_to_sum=[
                                5, 6, 7, 8, 9], columns_to_average=[9])
                        elif empresa == 9001 or empresa == 310 or empresa == 320:
                            process_data(sheet, datos, 8, columns_to_sum=[
                                4, 5, 6, 7, 8], columns_to_average=[8])
                elif Opcion == 9:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[0])
                    if SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     6, 7, 8, 9, 10], columns_to_average=[0])
                    if SubOpcion == 2:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     7, 8, 9, 10, 11], columns_to_average=[0])
                    if SubOpcion == 3:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4, 5, 6, 7], columns_to_average=[0])
                    if SubOpcion == 4:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     7, 8, 9, 10, 11], columns_to_average=[0])
                elif Opcion == 11:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4, 5], columns_to_average=[0])
                elif Opcion == 12:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     6], columns_to_average=[0])
                elif Opcion == 13:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     2], columns_to_average=[0])
                elif Opcion == 14:
                    if SubOpcion == 0:
                        try:
                            if SubOpcion2 == 0 or SubOpcion2 == 1:
                                process_data(sheet, datos, 8, columns_to_sum=[
                                    7], columns_to_average=[0])
                        except:
                            pass
                elif Opcion == 16:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[])
                elif Opcion == 20:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4], columns_to_average=[0])
                    elif SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     9], columns_to_average=[0])
                elif Opcion == 26:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4], columns_to_average=[0])
                    elif SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     9], columns_to_average=[0])
                elif Opcion == 28:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     8, 9], columns_to_average=[0])
                elif Opcion == 29:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     5, 6], columns_to_average=[0])
                    elif SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     8], columns_to_average=[0])
                elif Opcion == 30:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     8], columns_to_average=[0])
                elif Opcion == 31:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[0])
                elif Opcion == 33:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[0])
                    if SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[0])
                elif Opcion == 34:
                    if SubOpcion == 0:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4, 5, 6], columns_to_average=[0])
                    if SubOpcion == 1:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     9, 10], columns_to_average=[0])
                    if SubOpcion == 2:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     4], columns_to_average=[0])
                    if SubOpcion == 3:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     0], columns_to_average=[0])
                    if SubOpcion == 4:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     5, 6, 7], columns_to_average=[0])
                    if SubOpcion == 5:
                        process_data(sheet, datos, 8, columns_to_sum=[
                                     10, 11, 18], columns_to_average=[0])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})

        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})

    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxPD_GetExtractoTER(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            print(datos)
            empresa = 277
            startDate = results['startDate']
            startDate = datetime.strptime(
                startDate, "%Y-%m-%dT%H:%M:%S.%fZ")
            year = startDate.year
            month = startDate.month

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            plantilla_path = os.path.join(
                script_dir, '../docs/Plantillas/Plantilla_Rpto_Contabilidad_Directivo.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                # column_colEmp = 'C'
                # row_colEmp = 2
                # write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'F'
                row_colMes = 6
                column_colYear = 'H'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)
                print("After")
                process_data(sheet, datos, 8, columns_to_sum=[],
                             columns_to_average=[])
                print("Before")

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})


@csrf_exempt
@require_http_methods(["POST"])
def XlsxAC_ComTaqNoNomina(request):
    if request.method == 'POST':
        # Verificar si hay datos en la solicitud
        if request.body:
            # Decodificar los datos JSON de la solicitud
            results = json.loads(request.body)

            datos = json.loads(results['results'])
            empresa = results['empresa']
            month = results['month']
            year = results['year']

            # Obtener la ruta absoluta de la plantilla Excel en el mismo directorio que el script
            script_dir = os.path.dirname(__file__)

            plantilla_path = os.path.join(
                script_dir, '../docs/Plantillas/Plantilla_Rpto_Comisiones_Taquilleros_NoNomina.xlsx')

            # Verificar si existe el archivo
            if os.path.exists(plantilla_path):
                workbook = load_workbook(plantilla_path)
                sheet = workbook.active

                # Ubicación para el nombre de la empresa
                column_colEmp = 'C'
                row_colEmp = 2
                write_name_company(sheet, empresa, column_colEmp, row_colEmp)

                # Ubicación para month y year del reporte
                column_colMes = 'I'
                row_colMes = 6
                column_colYear = 'G'
                row_colYear = 6
                write_month_year(sheet, month, year, column_colMes,
                                 row_colMes, column_colYear, row_colYear)

                process_data(sheet, datos, 8, columns_to_sum=[],
                             columns_to_average=[])

                return generate_excel_response(workbook)
            else:
                return JsonResponse({'error': 'El archivo de la plantilla no fue encontrado'})
        else:
            return JsonResponse({'error': 'Se esperaba una solicitud POST'})
    else:
        return JsonResponse({'error': 'Se esperaba una solicitud POST'})
